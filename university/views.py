from django.shortcuts import render

# Create your views here.
from django.shortcuts import render, redirect,get_object_or_404
import pandas as pd
import openpyxl, os
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import JsonResponse, HttpResponse, FileResponse
from university.forms import LoginForm, StudentForm, MyclassForm
from datetime import datetime
from django.shortcuts import render
from django.http import HttpResponse
from datetime import date
from .models import Myclass
from .forms import MonthYearForm, UploadFileForm, UserCreationForm
from collections import defaultdict
from university.models import Student
from datetime import datetime, timedelta 
from university import models
import calendar
from django.template.loader import get_template
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User, Group
from .forms import UserRegisterForm
#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
@login_required
def home(request):
    return render(request, 'home.html')

@user_passes_test(lambda u: u.is_superuser or  Group.objects.get_or_create(name='YourGroupName')[0] in u.groups.all())
def group_specific_view(request):
    return render(request, 'group_specific.html')


def register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            group, created = Group.objects.get_or_create(name='YourGroupName')  # Replace with your group name
            user.groups.add(group)
            user.save()
            raw_password = form.cleaned_data.get('password1')
            user = authenticate(username=user.username, password=raw_password)
            login(request, user)
            return redirect('home')
    else:
        form = UserRegisterForm()
    return render(request, 'signup.html', {'form': form})


def upload_student_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            missing_details = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, roll_number, student_class, section = row
                if not all([name, roll_number, student_class, section]):
                    missing_details.append(row)
                else:
                    Student.objects.create(
                        name=name,
                        roll_number=roll_number,
                        student_class=student_class,
                        section=section
                    )

            if missing_details:
                return render(request, 'upload_students.html', {
                    'form': form,
                    'missing_details': missing_details
                })

            return redirect('success')

    else:
        form = UploadFileForm()
    
    return render(request, 'upload_students.html', {'form': form})

def success(request):
    return HttpResponse("Students uploaded successfully!")

def generate_individual_reports(request):
    students = Student.objects.all()
    for student in students:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Student Report'

        # Set up headers
        sheet['A1'] = 'Student Report'
        sheet['A1'].font = Font(size=14, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet.merge_cells('A1:D1')

        sheet['A2'] = 'Name:'
        sheet['B2'] = student.name
        sheet['A3'] = 'Roll Number:'
        sheet['B3'] = student.roll_number
        sheet['A4'] = 'Class:'
        sheet['B4'] = student.student_class
        sheet['A5'] = 'Section:'
        sheet['B5'] = student.section

        for col_num in range(1, 5):
            sheet.column_dimensions[get_column_letter(col_num)].width = 20

        # Save each report
        file_name = f'student_report_{student.roll_number}.xlsx'
        workbook.save(file_name)

    return HttpResponse("Reports generated successfully!")
def download_template(request):
    file_path = os.path.join('C:/Users/welcome/Desktop/project/college/university/templates/', 'student_template.xlsx')
    return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='student_template.xlsx')
    
@login_required
def select_class(request):
    classes = upload_student_file.objects.all()
    return render(request, 'select_class.html', {'classes': classes})

@login_required
def view1(request):
    classes = Student.objects.values_list('student_class', flat=True).distinct()
    selected_class = request.GET.get('class')
    sections = Student.objects.filter(student_class=selected_class).values_list('section', flat=True).distinct() if selected_class else []
    selected_section = request.GET.get('section')
    
    students = Student.objects.filter(student_class=selected_class, section=selected_section) if selected_class and selected_section else Student.objects.none()

    if request.method == 'POST':
        form_data = []
        for student in students:
            status = request.POST.get(f'status_{student.id}')
            if status:
                form_data.append({'student': student.id, 'status': status})
        
        # Create a form instance for each student to validate and save data
        for data in form_data:
            form = MyclassForm(data)
            if form.is_valid():
                form.save()
            else:
                return JsonResponse({'error': 'Form is not valid', 'details': form.errors}, status=400)
        
        return JsonResponse({'message': 'Attendance marked successfully'})
    
    else:
        form = MyclassForm()
    
    return render(request, 'Attendance.html', {
        'form': form,
        'students': students,
        'classes': classes,
        'selected_class': selected_class,
        'sections': sections,
        'selected_section': selected_section
    })


def view2(request):
    
    if request.method == 'POST':
        form = MonthYearForm(request.POST)
        if form.is_valid():
            month = form.cleaned_data['month']
            year = form.cleaned_data['year']

            # Create a new workbook and select the active sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Attendance Record'

            # Set up the headers
            sheet['A1'] = 'ATTENDANCE RECORD'
            sheet['A1'].font = Font(size=14, bold=True)
            sheet['A1'].alignment = Alignment(horizontal='center')
            sheet.merge_cells('A1:H1')

            sheet['B2'] = 'Month:'
            sheet['C2'] = datetime(year, month, 1).strftime('%B')
            sheet['B3'] = 'Year:'
            sheet['C3'] = year

            sheet['B2'].alignment = Alignment(horizontal='right')
            sheet['B3'].alignment = Alignment(horizontal='right')

            # Set column headers
            headers = ['Em Name', 'Roll Number', 'Class', 'Section']
            for day in range(1, 32):
                try:
                    current_date = datetime(year, month, day)
                    headers.append(current_date.strftime('%a %d'))
                except ValueError:
                    break
            headers.append('Total Days Present')

            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=5, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                if col_num > 4:
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            # Fill in the attendance data
            students = Student.objects.all()
            for row_num, student in enumerate(students, 6):
                sheet.cell(row=row_num, column=1).value = student.name
                sheet.cell(row=row_num, column=2).value = student.roll_number
                sheet.cell(row=row_num, column=3).value = student.student_class
                sheet.cell(row=row_num, column=4).value = student.section
                total_days_present = 0
                for day in range(1, 32):
                    try:
                        current_date = datetime(year, month, day)
                        attendance = Myclass.objects.filter(student=student, date=current_date).first()
                        if attendance:
                            status = attendance.status
                            if status in ['present', 'od']:
                                total_days_present += 1
                        else:
                            status = ''
                        cell = sheet.cell(row=row_num, column=day+4)
                        cell.value = status
                        if current_date.date() == date.today():
                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    except ValueError:
                        break
                # Write the total number of days present in the last column
                sheet.cell(row=row_num, column=len(headers)).value = total_days_present

            # Adjust column widths
            for col_num in range(1, len(headers) + 1):
                sheet.column_dimensions[get_column_letter(col_num)].width = 15

            # Save the workbook to a HttpResponse
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=attendance_{year}_{month}.xlsx'
            workbook.save(response)
            return response
    else:
        form = MonthYearForm()

    return render(request, 'generate_report.html', {'form': form})
